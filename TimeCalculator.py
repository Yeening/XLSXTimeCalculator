#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import sys

RECORD_COLUMN = 10

def calculate(value):
    timeSum = 0
    for time in value.strip().split("\n"):
        time = time.strip()
        startT = time.split('-')[0].strip()
        endT = time.split('-')[1].strip()
        startM, startS = int(startT[:-3]), int(startT[-2:])
        endM, endS = int(endT[:-3]), int(endT[-2:])
        # print(startM,endM)
        timeSum += (endM - startM) * 60 + (endS - startS)
    return timeSum

def isTime(s):
    if s is None:
        return False
    if(s.split("\n")) is None:
        return False
    return s.strip().split("\n")[0].split('-')[0].isnumeric() or s.strip().split("\n")[0].split(':')[0].isnumeric()


if __name__ == '__main__':
    if(len(sys.argv)>1):
        file_path = sys.argv[1]
    else:
        file_path = 'EP9.xlsx'
    print("输入文件：",file_path)
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    # g6 = sheet['J15']
    for i in range(5, 17):
        curTime = 0
        for j in range(1,10):
            cell = sheet.cell(row = i, column=j).value
            if(isTime(cell)):
                curTime += calculate(cell)
        sheet.cell(row = i, column=RECORD_COLUMN).value = str(curTime)
    book.save('result.xlsx')
    print("输出文件：", 'result.xlsx')
